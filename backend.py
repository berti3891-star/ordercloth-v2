from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import io
from datetime import datetime
import json
import os
from functools import wraps

app = Flask(__name__)
CORS(app)

# ============================================================
# CONFIGURAZIONE SICUREZZA A 2 LIVELLI
# ============================================================

SUPER_ADMIN = {
    'email': 'berti3891@gmail.com',
    'password': 'N\'F8g3gq+KM0',
    'role': 'super_admin'
}

MANAGER = {
    'username': 'manager',
    'password': 'manager123',
    'role': 'manager'
}

# Stato globale dell'app
app_status = {
    'enabled': True,
    'disabled_at': None,
    'disabled_by': None
}

# Configurazione Gmail
GMAIL_ADDRESS = 'berti3891@gmail.com'
GMAIL_PASSWORD = 'lwuueawecubwybyv'

# ============================================================
# DECORATORI AUTENTICAZIONE
# ============================================================

def require_super_admin(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        token = request.headers.get('Authorization')
        
        if not token:
            return jsonify({'error': 'Token mancante'}), 401
        
        # Verifica token (semplice implementazione)
        try:
            auth_data = json.loads(token)
            if auth_data.get('role') != 'super_admin':
                return jsonify({'error': 'Accesso negato - Solo Super Admin'}), 403
        except:
            return jsonify({'error': 'Token invalido'}), 401
        
        return f(*args, **kwargs)
    return decorated_function

def require_manager(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        token = request.headers.get('Authorization')
        
        if not token:
            return jsonify({'error': 'Token mancante'}), 401
        
        try:
            auth_data = json.loads(token)
            if auth_data.get('role') not in ['super_admin', 'manager']:
                return jsonify({'error': 'Accesso negato'}), 403
        except:
            return jsonify({'error': 'Token invalido'}), 401
        
        return f(*args, **kwargs)
    return decorated_function

def check_app_enabled(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not app_status['enabled']:
            return jsonify({
                'error': 'App disabilitata',
                'disabled_at': app_status['disabled_at'],
                'message': 'L\'app è stata disabilitata da Super Admin'
            }), 503
        
        return f(*args, **kwargs)
    return decorated_function

# ============================================================
# ENDPOINTS AUTENTICAZIONE
# ============================================================

@app.route('/api/auth/login', methods=['POST'])
def login():
    """
    Endpoint per il login
    Body: { "email/username": "...", "password": "..." }
    """
    data = request.json
    identifier = data.get('identifier')
    password = data.get('password')
    
    if not identifier or not password:
        return jsonify({'error': 'Identifier e password richiesti'}), 400
    
    # Controlla Super Admin
    if identifier == SUPER_ADMIN['email'] and password == SUPER_ADMIN['password']:
        return jsonify({
            'success': True,
            'user': {
                'email': SUPER_ADMIN['email'],
                'role': 'super_admin',
                'permissions': ['view_all', 'edit_all', 'manage_users', 'disable_app']
            },
            'token': json.dumps({'role': 'super_admin', 'email': SUPER_ADMIN['email']})
        }), 200
    
    # Controlla Manager
    if identifier == MANAGER['username'] and password == MANAGER['password']:
        return jsonify({
            'success': True,
            'user': {
                'username': MANAGER['username'],
                'role': 'manager',
                'permissions': ['edit_catalogo']
            },
            'token': json.dumps({'role': 'manager', 'username': MANAGER['username']})
        }), 200
    
    return jsonify({'error': 'Credenziali non valide'}), 401

@app.route('/api/auth/status', methods=['GET'])
def auth_status():
    """
    Endpoint per verificare lo stato dell'autenticazione
    """
    return jsonify({
        'app_enabled': app_status['enabled'],
        'disabled_at': app_status['disabled_at']
    }), 200

# ============================================================
# ENDPOINTS APP STATUS
# ============================================================

@app.route('/api/app/status', methods=['GET'])
def get_app_status():
    """
    Verifica se l'app è abilitata o disabilitata
    """
    return jsonify({
        'enabled': app_status['enabled'],
        'disabled_at': app_status['disabled_at'],
        'disabled_by': app_status['disabled_by']
    }), 200

@app.route('/api/app/disable', methods=['POST'])
@require_super_admin
def disable_app():
    """
    Super Admin disabilita tutte le app
    """
    global app_status
    
    data = request.json
    reason = data.get('reason', 'Disabilitata da Super Admin')
    
    app_status['enabled'] = False
    app_status['disabled_at'] = datetime.now().isoformat()
    app_status['disabled_by'] = SUPER_ADMIN['email']
    
    # Log della disabilitazione
    print(f"[SECURITY] App disabilitata da {SUPER_ADMIN['email']} - {reason}")
    
    return jsonify({
        'success': True,
        'message': 'App disabilitata',
        'disabled_at': app_status['disabled_at']
    }), 200

@app.route('/api/app/enable', methods=['POST'])
@require_super_admin
def enable_app():
    """
    Super Admin abilita le app
    """
    global app_status
    
    app_status['enabled'] = True
    app_status['disabled_at'] = None
    
    print(f"[SECURITY] App riabilitata da {SUPER_ADMIN['email']}")
    
    return jsonify({
        'success': True,
        'message': 'App abilitata'
    }), 200

# ============================================================
# ENDPOINTS SECURITY
# ============================================================

@app.route('/api/admin/change-password', methods=['POST'])
@require_super_admin
def change_super_admin_password():
    """
    Super Admin cambia la propria password
    """
    global SUPER_ADMIN
    
    data = request.json
    new_password = data.get('new_password')
    
    if not new_password or len(new_password) < 4:
        return jsonify({'error': 'Password non valida (min 4 caratteri)'}), 400
    
    old_password = SUPER_ADMIN['password']
    SUPER_ADMIN['password'] = new_password
    
    print(f"[SECURITY] Password Super Admin cambiata")
    
    return jsonify({
        'success': True,
        'message': 'Password aggiornata'
    }), 200

@app.route('/api/manager/change-password', methods=['POST'])
@require_super_admin
def change_manager_password():
    """
    Super Admin cambia la password del Manager
    """
    global MANAGER
    
    data = request.json
    new_password = data.get('new_password')
    
    if not new_password or len(new_password) < 4:
        return jsonify({'error': 'Password non valida (min 4 caratteri)'}), 400
    
    MANAGER['password'] = new_password
    
    print(f"[SECURITY] Password Manager cambiata da Super Admin")
    
    return jsonify({
        'success': True,
        'message': 'Password Manager aggiornata'
    }), 200

# ============================================================
# ENDPOINTS CATALOGO (Manager + Super Admin)
# ============================================================

@app.route('/api/catalogo/update', methods=['POST'])
@require_manager
def update_catalogo():
    """
    Manager (o Super Admin) modifica il catalogo
    """
    data = request.json
    
    # Qui salvare le modifiche nel database Excel
    print(f"[CATALOGO] Catalogo aggiornato")
    
    return jsonify({
        'success': True,
        'message': 'Catalogo aggiornato'
    }), 200

# ============================================================
# ENDPOINTS ORDINI
# ============================================================

@app.route('/api/ordine', methods=['POST'])
@check_app_enabled
def crea_ordine():
    """
    Crea un nuovo ordine
    """
    try:
        dati_ordine = request.json
        
        if not dati_ordine.get('ragioneSociale') or not dati_ordine.get('articoli'):
            return jsonify({'error': 'Dati mancanti'}), 400
        
        # Converti il totale
        totale_str = dati_ordine.get('totale', '€ 0.00').replace('€ ', '').replace(',', '.')
        totale = float(totale_str)
        dati_ordine['totale'] = totale
        
        # Crea il file Excel
        file_excel = crea_excel_ordine(dati_ordine)
        nome_file = f"Ordine_{dati_ordine['piva']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Prepara gli indirizzi email
        email_cliente = dati_ordine.get('emailCliente')
        email_rappresentante = dati_ordine.get('emailRappresentante')
        
        # Invia email
        if email_cliente:
            file_excel.seek(0)
            corpo_cliente = formatta_email_cliente(dati_ordine)
            invia_email(
                email_cliente,
                f"✅ Ordine Confermato - {dati_ordine['ragioneSociale']}",
                corpo_cliente,
                file_excel,
                nome_file
            )
        
        if email_rappresentante:
            file_excel.seek(0)
            corpo_cliente = formatta_email_cliente(dati_ordine)
            invia_email(
                email_rappresentante,
                f"✅ Ordine Confermato - {dati_ordine['ragioneSociale']}",
                corpo_cliente,
                file_excel,
                nome_file
            )
        
        # Invia email all'admin
        file_excel.seek(0)
        corpo_admin = formatta_email_admin(dati_ordine)
        invia_email(
            GMAIL_ADDRESS,
            f"📦 Nuovo Ordine - {dati_ordine['ragioneSociale']}",
            corpo_admin,
            file_excel,
            nome_file
        )
        
        return jsonify({
            'success': True,
            'message': 'Ordine processato con successo',
            'ordine_id': nome_file
        }), 200
    
    except Exception as e:
        print(f"Errore: {str(e)}")
        return jsonify({'error': str(e)}), 500

# ============================================================
# FUNZIONI UTILITÀ
# ============================================================

def crea_excel_ordine(dati_ordine):
    """Crea file Excel con i dati dell'ordine"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ordine"
    
    header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14, color="1F4788")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    ws['A1'] = "ORDINE ABBIGLIAMENTO"
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = center_alignment
    
    ws['A2'] = f"Data: {dati_ordine['data']}"
    ws['A3'] = f"Catalogo: {dati_ordine['catalogo'].upper()}"
    
    ws['A5'] = "DATI CLIENTE"
    ws['A5'].font = Font(bold=True, color="1F4788")
    
    row = 6
    ws[f'A{row}'] = "Ragione Sociale:"
    ws[f'B{row}'] = dati_ordine['ragioneSociale']
    row += 1
    
    ws[f'A{row}'] = "P.IVA:"
    ws[f'B{row}'] = dati_ordine['piva']
    row += 1
    
    ws[f'A{row}'] = "Indirizzo Consegna:"
    ws[f'B{row}'] = dati_ordine['indirizzo']
    
    row = 12
    ws[f'A{row}'] = "ARTICOLI ORDINATI"
    ws[f'A{row}'].font = Font(bold=True, color="1F4788")
    
    row += 1
    headers = ["Articolo", "Colore", "Taglia", "Quantità", "Prezzo Unitario", "Totale"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    row += 1
    for articolo in dati_ordine['articoli']:
        ws.cell(row=row, column=1).value = articolo['nome']
        ws.cell(row=row, column=2).value = articolo['colore']
        ws.cell(row=row, column=3).value = articolo['taglia']
        ws.cell(row=row, column=4).value = articolo['quantita']
        ws.cell(row=row, column=5).value = articolo['prezzoUnitario']
        ws.cell(row=row, column=6).value = articolo['prezzoTotale']
        
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if col in [5, 6]:
                cell.number_format = '€ #,##0.00'
        
        row += 1
    
    row += 1
    ws[f'E{row}'] = "TOTALE ORDINE:"
    ws[f'E{row}'].font = Font(bold=True, size=12, color="1F4788")
    ws[f'F{row}'] = dati_ordine['totale']
    ws[f'F{row}'].font = Font(bold=True, size=12, color="1F4788")
    ws[f'F{row}'].number_format = '€ #,##0.00'
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    return file_stream

def invia_email(destinatari, oggetto, corpo, file_excel=None, nome_file=None):
    """Invia email via Gmail con allegato Excel"""
    try:
        msg = MIMEMultipart()
        msg['From'] = GMAIL_ADDRESS
        msg['To'] = destinatari
        msg['Subject'] = oggetto
        
        msg.attach(MIMEText(corpo, 'html', 'utf-8'))
        
        if file_excel:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(file_excel.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename= {nome_file}')
            msg.attach(part)
        
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(GMAIL_ADDRESS, GMAIL_PASSWORD)
        server.send_message(msg)
        server.quit()
        
        return True
    
    except Exception as e:
        print(f"Errore invio email: {str(e)}")
        return False

def formatta_email_cliente(dati_ordine):
    """Formatta l'email per il cliente"""
    html = f"""
    <html>
        <body style="font-family: Arial, sans-serif; color: #333;">
            <h2 style="color: #1F4788;">✅ Ordine Confermato</h2>
            <p>Caro cliente,</p>
            <p>Ti confermiamo la ricezione del tuo ordine.</p>
            
            <h3 style="color: #1F4788;">Dettagli Ordine:</h3>
            <p><strong>Data:</strong> {dati_ordine['data']}</p>
            <p><strong>Catalogo:</strong> {dati_ordine['catalogo'].upper()}</p>
            <p><strong>Totale:</strong> {dati_ordine['totale']}</p>
            
            <p style="margin-top: 20px; color: #666; font-size: 12px;">
                L'ordine è in allegato in formato Excel.
            </p>
        </body>
    </html>
    """
    return html

def formatta_email_admin(dati_ordine):
    """Formatta l'email per l'admin"""
    html = f"""
    <html>
        <body style="font-family: Arial, sans-serif; color: #333;">
            <h2 style="color: #1F4788;">📦 Nuovo Ordine Ricevuto</h2>
            
            <h3 style="color: #1F4788;">Informazioni Cliente:</h3>
            <p><strong>Ragione Sociale:</strong> {dati_ordine['ragioneSociale']}</p>
            <p><strong>P.IVA:</strong> {dati_ordine['piva']}</p>
            <p><strong>Email:</strong> {dati_ordine['emailCliente']}</p>
            
            <h3 style="color: #1F4788;">Dettagli Ordine:</h3>
            <p><strong>Data:</strong> {dati_ordine['data']}</p>
            <p><strong>Catalogo:</strong> {dati_ordine['catalogo'].upper()}</p>
            <p><strong>Totale:</strong> {dati_ordine['totale']}</p>
        </body>
    </html>
    """
    return html

# ============================================================
# HEALTH CHECK
# ============================================================

@app.route('/api/health', methods=['GET'])
def health():
    """Health check endpoint"""
    return jsonify({
        'status': 'ok',
        'app_enabled': app_status['enabled']
    }), 200

# ============================================================
# MAIN
# ============================================================

if __name__ == '__main__':
    print("""
    ========================================
    🚀 ORDER CLOTH v2.0 - BACKEND FLASK
    ========================================
    
    ⚠️  CONFIGURAZIONE:
    1. Configura GMAIL_PASSWORD
    2. Genera App Password: https://myaccount.google.com/apppasswords
    3. Copia in GMAIL_PASSWORD
    
    🔐 SICUREZZA A 2 LIVELLI:
    - Super Admin: berti3891@gmail.com
    - Manager: manager (username)
    - Blocco globale app: ✅ Implementato
    
    ========================================
    """)
    
    app.run(debug=True, host='0.0.0.0', port=5000)
